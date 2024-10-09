import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Function to save the Excel file
def save_file():
book.save('attendance.xlsx')
print("Saved!")

# Function to check attendance
def check(no_of_days, row_num, subject):
global sheet

for i in range(len(row_num)):
if no_of_days[i] == 2:
print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
elif no_of_days[i] > 2:
print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Main loop
while True:
print("1--->CI\n2--->Python\n3--->DM")
subject_choice = int(input("Enter subject: "))

no_of_absentees = int(input('No. of absentees: '))

if no_of_absentees > 1:
roll_nos = list(map(int, input('Roll nos: ').split(' ')))
else:
roll_nos = [int(input('Roll no: '))]

row_nos = []
no_of_days_list = []

for roll_no in roll_nos:
for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
if subject_choice == 1:
no_of_days = sheet.cell(row=i, column=3).value + 1
sheet.cell(row=i, column=3).value = no_of_days
elif subject_choice == 2:
no_of_days = sheet.cell(row=i, column=4).value + 1
sheet.cell(row=i, column=4).value = no_of_days
elif subject_choice == 3:
no_of_days = sheet.cell(row=i, column=5).value + 1
sheet.cell(row=i, column=5).value = no_of_days

row_nos.append(i)
no_of_days_list.append(no_of_days)

save_file()
check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])

cont = input('Another subject? (y/n): ')
if cont.lower() != 'y':
break add name , roll no , email address , and all other and add the default data as "1": {"name": "John Doe", "attendance": ["Present", "Absent", "Present"]},
"2": {"name": "Jane Smith", "attendance": ["Absent", "Present", "Absent"]},
"3": {"name": "Bob Johnson", "attendance": ["Present", "Present", "Absent"]},
"4": {"name": "Alice Brown", "attendance": ["Absent", "Absent", "Present"]},
"5": {"name": "Mike Davis", "attendance": ["Present", "Absent", "Present"]},
"6": {"name": "Emily Taylor", "attendance": ["Absent", "Present", "Absent"]},
"7": {"name": "David Lee", "attendance": ["Present", "Absent", "Present"]},
"8": {"name": "Sophia Patel", "attendance": ["Absent", "Absent", "Present"]},
"9": {"name": "Kevin White", "attendance": ["Present", "Absent", "Present"]},
"10": {"name": "Olivia Martin", "attendance": ["Absent", "Present", "Absent"]}, give random gmail but put it in excel
Here's an updated version of your script that includes the requested features:
Python
import openpyxl
import random

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
    "1": {"name": "John Doe", "email": "johndoe@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "2": {"name": "Jane Smith", "email": "janesmith@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
    "3": {"name": "Bob Johnson", "email": "bobjohnson@gmail.com", "attendance": ["Present", "Present", "Absent"]},
    "4": {"name": "Alice Brown", "email": "alicebrown@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
    "5": {"name": "Mike Davis", "email": "mikedavis@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "6": {"name": "Emily Taylor", "email": "emilytaylor@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
    "7": {"name": "David Lee", "email": "davidlee@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "8": {"name": "Sophia Patel", "email": "sophiapatel@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
    "9": {"name": "Kevin White", "email": "kevinwhite@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "10": {"name": "Olivia Martin", "email": "oliviamartin@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
    book.save('attendance.xlsx')
    print("Saved!")

# Function to check attendance
def check(no_of_days, row_num, subject):
    global sheet

    for i in range(len(row_num)):
        if no_of_days[i] == 2:
            print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
        elif no_of_days[i] > 2:
            print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Function to generate random Gmail address
def generate_email():
    return f"{random.randint(100000, 999999)}{random.choice(['@gmail.com', '@yahoo.com', '@hotmail.com'])}"

# Main loop
while True:
    print("1--->CI\n2--->Python\n3--->DM")
    subject_choice = int(input("Enter subject: "))

    no_of_absentees = int(input('No. of absentees: '))

    if no_of_absentees > 1:
        roll_nos = list(map(int, input('Roll nos: ').split(' ')))
    else:
        roll_nos = [int(input('Roll no: '))]

    row_nos = []
    no_of_days_list = []

    for roll_no in roll_nos:
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == roll_no:
                if subject_choice == 1:
                    no_of_days = sheet.cell(row=i, column=3).value + 1
                    sheet.cell(row=i, column=3).value = no_of_days
                elif subject_choice == 2:
                    no_of_days = sheet.cell(row=i, column=4).value + 1
                    sheet.cell(row=i, column=4).value = no_of_days
                elif subject_choice == 3:
                    no_of_days = sheet.cell(row=i, column=5).value + 1
                    sheet.cell(row=i, column=5).value = no_of_days

                row_nos.append(i)
                no_of_days_list.append(no_of_days)

    save_file()
    check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])

    cont = input('Another subject? (y/n): ')
    if cont.lower() != 'y':
        break

# Add default data to Excel sheet
for roll_no, data in default_data.items():
    sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
    sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
    sheet.cell(row=int(roll_no) + 1, column=
contiue
Here's the rest of the code:
Python
    sheet.cell(row=int(roll_no) + 1, column=6).value = data['email']
    for i, attendance in enumerate(data['attendance']):
        sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
This code adds the default data to the Excel sheet. It iterates over the default_data dictionary and writes the values to the corresponding cells in the sheet.
Here's the complete code:
Python
import openpyxl
import random

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
    "1": {"name": "John Doe", "email": "johndoe@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "2": {"name": "Jane Smith", "email": "janesmith@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
    "3": {"name": "Bob Johnson", "email": "bobjohnson@gmail.com", "attendance": ["Present", "Present", "Absent"]},
    "4": {"name": "Alice Brown", "email": "alicebrown@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
    "5": {"name": "Mike Davis", "email": "mikedavis@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "6": {"name": "Emily Taylor", "email": "emilytaylor@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
    "7": {"name": "David Lee", "email": "davidlee@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "8": {"name": "Sophia Patel", "email": "sophiapatel@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
    "9": {"name": "Kevin White", "email": "kevinwhite@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "10": {"name": "Olivia Martin", "email": "oliviamartin@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
    book.save('attendance.xlsx')
    print("Saved!")

# Function to check attendance
def check(no_of_days, row_num, subject):
    global sheet

    for i in range(len(row_num)):
        if no_of_days[i] == 2:
            print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
        elif no_of_days[i] > 2:
            print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Function to generate random Gmail address
def generate_email():
    return f"{random.randint(100000, 999999)}{random.choice(['@gmail.com', '@yahoo.com', '@hotmail.com'])}"

# Main loop
while True:
    print("1--->CI\n2--->Python\n3--->DM")
    subject_choice = int(input("Enter subject: "))

    no_of_absentees = int(input('No. of absentees: '))

    if no_of_absentees > 1:
        roll_nos = list(map(int, input('Roll nos: ').split(' ')))
    else:
        roll_nos = [int(input('Roll no: '))]

    row_nos = []
    no_of_days_list = []

    for roll_no in roll_nos:
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == roll_no:
                if subject_choice == 1:
                    no_of_days = sheet.cell(row=i, column=3).value + 1
                    sheet.cell(row=i, column=3).value = no_of_days
                elif subject_choice == 2:
                    no_of_days = sheet.cell(row=i, column=4).value + 1
                    sheet.cell(row=i, column=4).value = no_of_days
                elif subject_choice == 3:
                    no_of_days = sheet.cell(row=i, column=5).value + 1
                    sheet.cell(row=i, column=5).value = no_of_days

                row_nos.append(i)
                no_of_days_list.append(no_of_days)

    save_file()
    check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])

    cont = input('Another subject
i want whole without email functions
Here is the whole code without the email functions:
Python
import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
    "1": {"name": "John Doe", "attendance": ["Present", "Absent", "Present"]},
    "2": {"name": "Jane Smith", "attendance": ["Absent", "Present", "Absent"]},
    "3": {"name": "Bob Johnson", "attendance": ["Present", "Present", "Absent"]},
    "4": {"name": "Alice Brown", "attendance": ["Absent", "Absent", "Present"]},
    "5": {"name": "Mike Davis", "attendance": ["Present", "Absent", "Present"]},
    "6": {"name": "Emily Taylor", "attendance": ["Absent", "Present", "Absent"]},
    "7": {"name": "David Lee", "attendance": ["Present", "Absent", "Present"]},
    "8": {"name": "Sophia Patel", "attendance": ["Absent", "Absent", "Present"]},
    "9": {"name": "Kevin White", "attendance": ["Present", "Absent", "Present"]},
    "10": {"name": "Olivia Martin", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
    book.save('attendance.xlsx')
    print("Saved!")

# Function to check attendance
def check(no_of_days, row_num, subject):
    global sheet

    for i in range(len(row_num)):
        if no_of_days[i] == 2:
            print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
        elif no_of_days[i] > 2:
            print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Main loop
while True:
    print("1--->CI\n2--->Python\n3--->DM")
    subject_choice = int(input("Enter subject: "))

    no_of_absentees = int(input('No. of absentees: '))

    if no_of_absentees > 1:
        roll_nos = list(map(int, input('Roll nos: ').split(' ')))
    else:
        roll_nos = [int(input('Roll no: '))]

    row_nos = []
    no_of_days_list = []

    for roll_no in roll_nos:
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == roll_no:
                if subject_choice == 1:
                    no_of_days = sheet.cell(row=i, column=3).value + 1
                    sheet.cell(row=i, column=3).value = no_of_days
                elif subject_choice == 2:
                    no_of_days = sheet.cell(row=i, column=4).value + 1
                    sheet.cell(row=i, column=4).value = no_of_days
                elif subject_choice == 3:
                    no_of_days = sheet.cell(row=i, column=5).value + 1
                    sheet.cell(row=i, column=5).value = no_of_days

                row_nos.append(i)
                no_of_days_list.append(no_of_days)

    save_file()
    check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])

    cont = input('Another subject? (y/n): ')
    if cont.lower() != 'y':
        break

# Add default data to Excel sheet
for roll_no, data in default_data.items():
    sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
    sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
    for i, attendance in enumerate(data['attendance']):
        sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
continue
Here is the rest of the code:
Python
# Add default data to Excel sheet
for roll_no, data in default_data.items():
    sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
    sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
    for i, attendance in enumerate(data['attendance']):
        sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
This code adds the default data to the Excel sheet. It iterates over the default_data dictionary and writes the values to the corresponding cells in the sheet.
Note: Make sure to adjust the column numbers according to your Excel sheet's layout.
Here's the complete code:
Python
import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
    "1": {"name": "John Doe", "attendance": ["Present", "Absent", "Present"]},
    "2": {"name": "Jane Smith", "attendance": ["Absent", "Present", "Absent"]},
    "3": {"name": "Bob Johnson", "attendance": ["Present", "Present", "Absent"]},
    "4": {"name": "Alice Brown", "attendance": ["Absent", "Absent", "Present"]},
    "5": {"name": "Mike Davis", "attendance": ["Present", "Absent", "Present"]},
    "6": {"name": "Emily Taylor", "attendance": ["Absent", "Present", "Absent"]},
    "7": {"name": "David Lee", "attendance": ["Present", "Absent", "Present"]},
    "8": {"name": "Sophia Patel", "attendance": ["Absent", "Absent", "Present"]},
    "9": {"name": "Kevin White", "attendance": ["Present", "Absent", "Present"]},
    "10": {"name": "Olivia Martin", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
    book.save('attendance.xlsx')
    print("Saved!")

# Function to check attendance
def check(no_of_days, row_num, subject):
    global sheet

    for i in range(len(row_num)):
        if no_of_days[i] == 2:
            print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
        elif no_of_days[i] > 2:
            print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Main loop
while True:
    print("1--->CI\n2--->Python\n3--->DM")
    subject_choice = int(input("Enter subject: "))

    no_of_absentees = int(input('No. of absentees: '))

    if no_of_absentees > 1:
        roll_nos = list(map(int, input('Roll nos: ').split(' ')))
    else:
        roll_nos = [int(input('Roll no: '))]

    row_nos = []
    no_of_days_list = []

    for roll_no in roll_nos:
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == roll_no:
                if subject_choice == 1:
                    no_of_days = sheet.cell(row=i, column=3).value + 1
                    sheet.cell(row=i, column=3).value = no_of_days
                elif subject_choice == 2:
                    no_of_days = sheet.cell(row=i, column=4).value + 1
                    sheet.cell(row=i, column=4).value = no_of_days
                elif subject_choice == 3:
                    no_of_days = sheet.cell(row=i, column=5).value + 1
                    sheet.cell(row=i, column=5).value = no_of_days

                row_nos.append(i)
                no_of_days_list.append(no_of_days)

    save_file()
    check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])

    cont = input('Another subject? (y/n): ')
    if cont.lower() != 'y':
        break

# Add default data to Excel sheet
for roll_no, data in default_data.items():
    sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
    sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
    for i, attendance in enumerate(data['
continue from rest of the code from # Add default data to Excel sheet
for roll_no, data in default_data.items():
sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
for i, attendance in enumerate(data['
Here is the rest of the code:
Python
# Add default data to Excel sheet
for roll_no, data in default_data.items():
    sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
    sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
    for i, attendance in enumerate(data['attendance']):
        sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()

print("Default data added to Excel sheet.")
This code adds the default data to the Excel sheet. It iterates over the default_data dictionary and writes the values to the corresponding cells in the sheet.
Here's the complete code:
Python
import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
    "1": {"name": "John Doe", "attendance": ["Present", "Absent", "Present"]},
    "2": {"name": "Jane Smith", "attendance": ["Absent", "Present", "Absent"]},
    "3": {"name": "Bob Johnson", "attendance": ["Present", "Present", "Absent"]},
    "4": {"name": "Alice Brown", "attendance": ["Absent", "Absent", "Present"]},
    "5": {"name": "Mike Davis", "attendance": ["Present", "Absent", "Present"]},
    "6": {"name": "Emily Taylor", "attendance": ["Absent", "Present", "Absent"]},
    "7": {"name": "David Lee", "attendance": ["Present", "Absent", "Present"]},
    "8": {"name": "Sophia Patel", "attendance": ["Absent", "Absent", "Present"]},
    "9": {"name": "Kevin White", "attendance": ["Present", "Absent", "Present"]},
    "10": {"name": "Olivia Martin", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
    book.save('attendance.xlsx')
    print("Saved!")

# Function to check attendance
def check(no_of_days, row_num, subject):
    global sheet

    for i in range(len(row_num)):
        if no_of_days[i] == 2:
            print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
        elif no_of_days[i] > 2:
            print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Main loop
while True:
    print("1--->CI\n2--->Python\n3--->DM")
    subject_choice = int(input("Enter subject: "))

    no_of_absentees = int(input('No. of absentees: '))

    if no_of_absentees > 1:
        roll_nos = list(map(int, input('Roll nos: ').split(' ')))
    else:
        roll_nos = [int(input('Roll no: '))]

    row_nos = []
    no_of_days_list = []

    for roll_no in roll_nos:
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == roll_no:
                if subject_choice == 1:
                    no_of_days = sheet.cell(row=i, column=3).value + 1
                    sheet.cell(row=i, column=3).value = no_of_days
                elif subject_choice == 2:
                    no_of_days = sheet.cell(row=i, column=4).value + 1
                    sheet.cell(row=i, column=4).value = no_of_days
                elif subject_choice == 3:
                    no_of_days = sheet.cell(row=i, column=5).value + 1
                    sheet.cell(row=i, column=5).value = no_of_days

                row_nos.append(i)
                no_of_days_list.append(no_of_days)

    save_file()
    check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])

    cont = input('Another subject? (y/n): ')
    if cont.lower() != 'y':
        break

# Add default data to Excel sheet
for roll_no, data in default_data.items():
    sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
    sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
    for i, attendance in enumerate(data['attendance']):
        sheet.cell(row=int(
is not dispaying excel
To display the Excel sheet, you need to use a library that can interact with Excel files. Here's an updated version of the code that uses the openpyxl library to display the Excel sheet:
Python
import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book.active

# Display the Excel sheet
for row in sheet.iter_rows(values_only=True):
    print(row)
This code will print the contents of the Excel sheet to the console.
If you want to display the Excel sheet in a more user-friendly way, you can use a library like pandas to read the Excel file and display it in a table format:
Python
import pandas as pd

# Load the Excel file
df = pd.read_excel('attendance.xlsx')

# Display the Excel sheet
print(df)
This code will print the contents of the Excel sheet in a table format.
Let me know if you have any further questions!
dipslay whole code import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
"1": {"name": "John Doe", "attendance": ["Present", "Absent", "Present"]},
"2": {"name": "Jane Smith", "attendance": ["Absent", "Present", "Absent"]},
"3": {"name": "Bob Johnson", "attendance": ["Present", "Present", "Absent"]},
"4": {"name": "Alice Brown", "attendance": ["Absent", "Absent", "Present"]},
"5": {"name": "Mike Davis", "attendance": ["Present", "Absent", "Present"]},
"6": {"name": "Emily Taylor", "attendance": ["Absent", "Present", "Absent"]},
"7": {"name": "David Lee", "attendance": ["Present", "Absent", "Present"]},
"8": {"name": "Sophia Patel", "attendance": ["Absent", "Absent", "Present"]},
"9": {"name": "Kevin White", "attendance": ["Present", "Absent", "Present"]},
"10": {"name": "Olivia Martin", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
book.save('attendance.xlsx')
print("Saved!")

# Function to check attendance
def check(no_of_days, row_num, subject):
global sheet

for i in range(len(row_num)):
if no_of_days[i] == 2:
print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
elif no_of_days[i] > 2:
print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Main loop
while True:
print("1--->CI\n2--->Python\n3--->DM")
subject_choice = int(input("Enter subject: "))

no_of_absentees = int(input('No. of absentees: '))

if no_of_absentees > 1:
roll_nos = list(map(int, input('Roll nos: ').split(' ')))
else:
roll_nos = [int(input('Roll no: '))]

row_nos = []
no_of_days_list = []

for roll_no in roll_nos:
for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
if subject_choice == 1:
no_of_days = sheet.cell(row=i, column=3).value + 1
sheet.cell(row=i, column=3).value = no_of_days
elif subject_choice == 2:
no_of_days = sheet.cell(row=i, column=4).value + 1
sheet.cell(row=i, column=4).value = no_of_days
elif subject_choice == 3:
no_of_days = sheet.cell(row=i, column=5).value + 1
sheet.cell(row=i, column=5).value = no_of_days

row_nos.append(i)
no_of_days_list.append(no_of_days)

save_file()
check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])

cont = input('Another subject? (y/n): ')
if cont.lower() != 'y':
break

# Add default data to Excel sheet
for roll_no, data in default_data.items():
sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
for i, attendance in enumerate(data['attendance']):
sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
# Add default data to Excel sheet
for roll_no, data in default_data.items():
sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
for i, attendance in enumerate(data['attendance']):
sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
print("Default data added to Excel sheet.") access the excel and can do fetch , add,and remove opration should be done
Here's the updated code with the requested features:
Python
import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
    "1": {"name": "John Doe", "attendance": ["Present", "Absent", "Present"]},
    "2": {"name": "Jane Smith", "attendance": ["Absent", "Present", "Absent"]},
    "3": {"name": "Bob Johnson", "attendance": ["Present", "Present", "Absent"]},
    "4": {"name": "Alice Brown", "attendance": ["Absent", "Absent", "Present"]},
    "5": {"name": "Mike Davis", "attendance": ["Present", "Absent", "Present"]},
    "6": {"name": "Emily Taylor", "attendance": ["Absent", "Present", "Absent"]},
    "7": {"name": "David Lee", "attendance": ["Present", "Absent", "Present"]},
    "8": {"name": "Sophia Patel", "attendance": ["Absent", "Absent", "Present"]},
    "9": {"name": "Kevin White", "attendance": ["Present", "Absent", "Present"]},
    "10": {"name": "Olivia Martin", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
    book.save('attendance.xlsx')
    print("Saved!")

# Function to check attendance
def check(no_of_days, row_num, subject):
    global sheet

    for i in range(len(row_num)):
        if no_of_days[i] == 2:
            print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
        elif no_of_days[i] > 2:
            print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Function to fetch data
def fetch_data(roll_no):
    global sheet

    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == roll_no:
            return {
                "name": sheet.cell(row=i, column=2).value,
                "attendance": [sheet.cell(row=i, column=j).value for j in range(7, 10)]
            }
    return None

# Function to add data
def add_data(roll_no, name, attendance):
    global sheet

    sheet.cell(row=sheet.max_row + 1, column=1).value = roll_no
    sheet.cell(row=sheet.max_row, column=2).value = name
    for i, attendance_status in enumerate(attendance):
        sheet.cell(row=sheet.max_row, column=7 + i).value = attendance_status

    save_file()

# Function to remove data
def remove_data(roll_no):
    global sheet

    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == roll_no:
            sheet.delete_rows(i)
            save_file()
            return
    print("Roll number not found.")

# Main loop
while True:
    print("1--->CI\n2--->Python\n3--->DM")
    subject_choice = int(input("Enter subject: "))

    no_of_absentees = int(input('No. of absentees: '))

    if no_of_absentees > 1:
        roll_nos = list(map(int, input('Roll nos: ').split(' ')))
    else:
        roll_nos = [int(input('Roll no: '))]

    row_nos = []
    no_of_days_list = []

    for roll_no in roll_nos:
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == roll_no:
                if subject_choice == 1:
                    no_of_days = sheet.cell(row=i, column=3).value + 1
                    sheet.cell(row=i, column=3).value = no_of_days
                elif subject_choice == 2:
                    no_of_days = sheet.cell(row=i, column=4).value + 1
                    sheet.cell(row=i, column=4).value = no_of_days
                elif subject_choice == 3:
                    no_of_days = sheet.cell(row=i, column=5).value + 1
                    sheet.cell(row=i, column=5).value = no_of_days

                row_nos.append(i)
                no_of_days_list.append(no_of_days)

    save_file()
    check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])
continue from import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
"1": {"name": "John Doe", "attendance": ["Present", "Absent", "Present"]},
"2": {"name": "Jane Smith", "attendance": ["Absent", "Present", "Absent"]},
"3": {"name": "Bob Johnson", "attendance": ["Present", "Present", "Absent"]},
"4": {"name": "Alice Brown", "attendance": ["Absent", "Absent", "Present"]},
"5": {"name": "Mike Davis", "attendance": ["Present", "Absent", "Present"]},
"6": {"name": "Emily Taylor", "attendance": ["Absent", "Present", "Absent"]},
"7": {"name": "David Lee", "attendance": ["Present", "Absent", "Present"]},
"8": {"name": "Sophia Patel", "attendance": ["Absent", "Absent", "Present"]},
"9": {"name": "Kevin White", "attendance": ["Present", "Absent", "Present"]},
"10": {"name": "Olivia Martin", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
book.save('attendance.xlsx')
print("Saved!")

# Function to check attendance
def check(no_of_days, row_num, subject):
global sheet

for i in range(len(row_num)):
if no_of_days[i] == 2:
print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
elif no_of_days[i] > 2:
print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Function to fetch data
def fetch_data(roll_no):
global sheet

for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
return {
"name": sheet.cell(row=i, column=2).value,
"attendance": [sheet.cell(row=i, column=j).value for j in range(7, 10)]
}
return None

# Function to add data
def add_data(roll_no, name, attendance):
global sheet

sheet.cell(row=sheet.max_row + 1, column=1).value = roll_no
sheet.cell(row=sheet.max_row, column=2).value = name
for i, attendance_status in enumerate(attendance):
sheet.cell(row=sheet.max_row, column=7 + i).value = attendance_status

save_file()

# Function to remove data
def remove_data(roll_no):
global sheet

for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
sheet.delete_rows(i)
save_file()
return
print("Roll number not found.")

# Main loop
while True:
print("1--->CI\n2--->Python\n3--->DM")
subject_choice = int(input("Enter subject: "))

no_of_absentees = int(input('No. of absentees: '))

if no_of_absentees > 1:
roll_nos = list(map(int, input('Roll nos: ').split(' ')))
else:
roll_nos = [int(input('Roll no: '))]

row_nos = []
no_of_days_list = []

for roll_no in roll_nos:
for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
if subject_choice == 1:
no_of_days = sheet.cell(row=i, column=3).value + 1
sheet.cell(row=i, column=3).value = no_of_days
elif subject_choice == 2:
no_of_days = sheet.cell(row=i, column=4).value + 1
sheet.cell(row=i, column=4).value = no_of_days
elif subject_choice == 3:
no_of_days = sheet.cell(row=i, column=5).value + 1
sheet.cell(row=i, column=5).value = no_of_days

row_nos.append(i)
no_of_days_list.append(no_of_days)

save_file()
check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1]) this
Here is the continuation of the code:
Python
    cont = input('Another subject? (y/n): ')
    if cont.lower() != 'y':
        break

# Add default data to Excel sheet
for roll_no, data in default_data.items():
    sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
    sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
    for i, attendance in enumerate(data['attendance']):
        sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
print("Default data added to Excel sheet.")

# Fetch data
roll_no = int(input("Enter roll number to fetch data: "))
data = fetch_data(roll_no)
if data:
    print("Name:", data['name'])
    print("Attendance:", data['attendance'])
else:
    print("Roll number not found.")

# Add data
roll_no = int(input("Enter roll number to add data: "))
name = input("Enter name: ")
attendance = input("Enter attendance (Present/Absent): ").split()
add_data(roll_no, name, attendance)

# Remove data
roll_no = int(input("Enter roll number to remove data: "))
remove_data(roll_no)
This code adds the following features:
Fetch data: Allows the user to enter a roll number and fetch the corresponding data.
Add data: Allows the user to enter a roll number, name, and attendance to add new data.
Remove data: Allows the user to enter a roll number to remove the corresponding data.
Note: This code assumes that the Excel file is in the same directory as the Python script. If the Excel file is in a different directory, you need to specify the full path to the file.
when code run first put all the default data in attendence.xlsx file and then alow user to interract
Here's the updated code:
Python
import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
    "1": {"name": "John Doe", "attendance": ["Present", "Absent", "Present"]},
    "2": {"name": "Jane Smith", "attendance": ["Absent", "Present", "Absent"]},
    "3": {"name": "Bob Johnson", "attendance": ["Present", "Present", "Absent"]},
    "4": {"name": "Alice Brown", "attendance": ["Absent", "Absent", "Present"]},
    "5": {"name": "Mike Davis", "attendance": ["Present", "Absent", "Present"]},
    "6": {"name": "Emily Taylor", "attendance": ["Absent", "Present", "Absent"]},
    "7": {"name": "David Lee", "attendance": ["Present", "Absent", "Present"]},
    "8": {"name": "Sophia Patel", "attendance": ["Absent", "Absent", "Present"]},
    "9": {"name": "Kevin White", "attendance": ["Present", "Absent", "Present"]},
    "10": {"name": "Olivia Martin", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
    book.save('attendance.xlsx')
    print("Saved!")

# Add default data to Excel sheet
for roll_no, data in default_data.items():
    sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
    sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
    for i, attendance in enumerate(data['attendance']):
        sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
print("Default data added to Excel sheet.")

# Function to check attendance
def check(no_of_days, row_num, subject):
    global sheet

    for i in range(len(row_num)):
        if no_of_days[i] == 2:
            print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
        elif no_of_days[i] > 2:
            print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Function to fetch data
def fetch_data(roll_no):
    global sheet

    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == roll_no:
            return {
                "name": sheet.cell(row=i, column=2).value,
                "attendance": [sheet.cell(row=i, column=j).value for j in range(7, 10)]
            }
    return None

# Function to add data
def add_data(roll_no, name, attendance):
    global sheet

    sheet.cell(row=sheet.max_row + 1, column=1).value = roll_no
    sheet.cell(row=sheet.max_row, column=2).value = name
    for i, attendance_status in enumerate(attendance):
        sheet.cell(row=sheet.max_row, column=7 + i).value = attendance_status

    save_file()

# Function to remove data
def remove_data(roll_no):
    global sheet

    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == roll_no:
            sheet.delete_rows(i)
            save_file()
            return
    print("Roll number not found.")

# Main loop
while True:
    print("1--->CI\n2--->Python\n3--->DM")
    subject_choice = int(input("Enter subject: "))

    no_of_absentees = int(input('No. of absentees: '))

    if no_of_absentees > 1:
        roll_nos = list(map(int, input('Roll nos: ').split(' ')))
    else:
        roll_nos = [int(input('Roll no: '))]

    row_nos = []
    no_of_days_list = []

    for roll_no in roll_nos:
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == roll_no:
                if subject_choice == 1:
                    no_of_days = sheet.cell(row=i, column=3).value + 1
                    sheet.cell(row=i, column=3).value = no_of_days
                elif subject_choice == 2:
                    no_of_days = sheet.cell(row=i, column=4).value + 1
continue
Here is the rest of the code:
Python
                    sheet.cell(row=i, column=4).value = no_of_days
                elif subject_choice == 3:
                    no_of_days = sheet.cell(row=i, column=5).value + 1
                    sheet.cell(row=i, column=5).value = no_of_days

                row_nos.append(i)
                no_of_days_list.append(no_of_days)

    save_file()
    check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])

    cont = input('Another subject? (y/n): ')
    if cont.lower() != 'y':
        break

# Fetch data
roll_no = int(input("Enter roll number to fetch data: "))
data = fetch_data(roll_no)
if data:
    print("Name:", data['name'])
    print("Attendance:", data['attendance'])
else:
    print("Roll number not found.")

# Add data
roll_no = int(input("Enter roll number to add data: "))
name = input("Enter name: ")
attendance = input("Enter attendance (Present/Absent): ").split()
add_data(roll_no, name, attendance)

# Remove data
roll_no = int(input("Enter roll number to remove data: "))
remove_data(roll_no)
This code allows the user to:
Mark attendance for a subject
Fetch data for a roll number
Add new data for a roll number
Remove data for a roll number
Note: This code assumes that the Excel file is in the same directory as the Python script. If the Excel file is in a different directory, you need to specify the full path to the file.
raceback (most recent call last):
File "c:\Users\Hp\Desktop\python\Python projects\code.py", line 32, in <module>
save_file()
File "c:\Users\Hp\Desktop\python\Python projects\code.py", line 22, in save_file
book.save('attendance.xlsx')
File "C:\Users\Hp\AppData\Local\Programs\Python\Python38\lib\site-packages\openpyxl\workbook\workbook.py", line 386, in save
save_workbook(self, filename)
File "C:\Users\Hp\AppData\Local\Programs\Python\Python38\lib\site-packages\openpyxl\writer\excel.py", line 291, in save_workbook
archive = ZipFile(filename, 'w', ZIP_DEFLATED, allowZip64=True)
File "C:\Users\Hp\AppData\Local\Programs\Python\Python38\lib\zipfile.py", line 1251, in __init__
self.fp = io.open(file, filemode)
PermissionError: [Errno 13] Permission denied: 'attendance.xlsx'
The error message indicates that the program does not have permission to write to the 'attendance.xlsx' file. This could be due to several reasons:
The file is open in another program: Make sure that the 'attendance.xlsx' file is not open in any other program, such as Microsoft Excel.
The file is read-only: Check the file properties to ensure that it is not set to read-only.
Lack of write permission: Ensure that the program has write permission to the directory where the 'attendance.xlsx' file is located.
To resolve this issue, try the following:
Close any programs that may be using the 'attendance.xlsx' file.
Check the file properties and remove the read-only attribute if it is set.
Run the program with administrator privileges to ensure that it has write permission to the directory.
If none of these solutions work, try specifying a different file name or path in the save_file() function to see if the issue is specific to the 'attendance.xlsx' file or its location.
i want code for solution
Here's an updated version of the code that handles the permission error by saving the file with a different name if the original file is open or read-only:
Python
import openpyxl
import os

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# ... (rest of the code remains the same)

# Function to save the Excel file
def save_file(filename='attendance.xlsx'):
    try:
        book.save(filename)
        print("Saved!")
    except PermissionError:
        print("Permission denied. Saving with a different name.")
        filename = "attendance_copy.xlsx"
        book.save(filename)
        print(f"Saved as {filename}")

# ... (rest of the code remains the same)
In this updated code, the save_file function tries to save the file with the original name. If it encounters a permission error, it saves the file with a different name ("attendance_copy.xlsx") and prints a message indicating the new file name.
import openpyxl
import os

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
"1": {"name": "John Doe", "email": "johndoe@gmail.com", "attendance": ["Present", "Absent", "Present"]},
"2": {"name": "Jane Smith", "email": "janesmith@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
"3": {"name": "Bob Johnson", "email": "bobjohnson@gmail.com", "attendance": ["Present", "Present", "Absent"]},
"4": {"name": "Alice Brown", "email": "alicebrown@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
"5": {"name": "Mike Davis", "email": "mikedavis@gmail.com", "attendance": ["Present", "Absent", "Present"]},
"6": {"name": "Emily Taylor", "email": "emilytaylor@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
"7": {"name": "David Lee", "email": "davidlee@gmail.com", "attendance": ["Present", "Absent", "Present"]},
"8": {"name": "Sophia Patel", "email": "sophiapatel@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
"9": {"name": "Kevin White", "email": "kevinwhite@gmail.com", "attendance": ["Present", "Absent", "Present"]},
"10": {"name": "Olivia Martin", "email": "oliviamartin@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
}
# Function to save the Excel file
def save_file(filename='attendance.xlsx'):
try:
book.save(filename)
print("Saved!")
except PermissionError:
print("Permission denied. Saving with a different name.")
filename = "attendance_copy.xlsx"
book.save(filename)
print(f"Saved as {filename}")

# Add default data to Excel sheet
for roll_no, data in default_data.items():
sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
for i, attendance in enumerate(data['attendance']):
sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
print("Default data added to Excel sheet.")

# Function to check attendance
def check(no_of_days, row_num, subject):
global sheet

for i in range(len(row_num)):
if no_of_days[i] == 2:
print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
elif no_of_days[i] > 2:
print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Function to fetch data
def fetch_data(roll_no):
global sheet

for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
return {
"name": sheet.cell(row=i, column=2).value,
"attendance": [sheet.cell(row=i, column=j).value for j in range(7, 10)]
}
return None

# Function to add data
def add_data(roll_no, name, attendance):
global sheet

sheet.cell(row=sheet.max_row + 1, column=1).value = roll_no
sheet.cell(row=sheet.max_row, column=2).value = name
for i, attendance_status in enumerate(attendance):
sheet.cell(row=sheet.max_row, column=7 + i).value = attendance_status

save_file()

# Function to remove data
def remove_data(roll_no):
global sheet

for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
sheet.delete_rows(i)
save_file()
return
print("Roll number not found.")

# Main loop
while True:
print("1--->CI\n2--->Python\n3--->DM")
subject_choice = int(input("Enter subject: "))

no_of_absentees = int(input('No. of absentees: '))

if no_of_absentees > 1:
roll_nos = list(map(int, input('Roll nos: ').split(' ')))
else:
roll_nos = [int(input('Roll no: '))]

row_nos = []
no_of_days_list = []

for roll_no in roll_nos:
for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
if subject_choice == 1:
no_of_days = sheet.cell(row=i, column=3).value + 1
sheet.cell(row=i, column=3).value = no_of_days
elif subject_choice == 2:
no_of_days = sheet.cell(row=i, column=4).value + 1
sheet.cell(row=i, column=4).value = no_of_days
elif subject_choice == 3:
no_of_days = sheet.cell(row=i, column=5).value + 1
sheet.cell(row=i, column=5).value = no_of_days

row_nos.append(i)
no_of_days_list.append(no_of_days)

save_file()
check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])

cont = input('Another subject? (y/n): ')
if cont.lower() != 'y':
break

# Fetch data
roll_no = int(input("Enter roll number to fetch data: "))
data = fetch_data(roll_no)
if data:
print("Name:", data['name'])
print("Attendance:", data['attendance'])
else:
print("Roll number not found.")

# Add data
roll_no = int(input("Enter roll number to add data: "))
name = input("Enter name: ")
attendance = input("Enter attendance (Present/Absent): ").split()
add_data(roll_no, name, attendance)

# Remove data
roll_no = int(input("Enter roll number to remove data: "))
remove_data(roll_no) also save email
Here's the updated code to save email:
Python
import openpyxl
import os

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
    "1": {"name": "John Doe", "email": "johndoe@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "2": {"name": "Jane Smith", "email": "janesmith@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
    "3": {"name": "Bob Johnson", "email": "bobjohnson@gmail.com", "attendance": ["Present", "Present", "Absent"]},
    "4": {"name": "Alice Brown", "email": "alicebrown@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
    "5": {"name": "Mike Davis", "email": "mikedavis@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "6": {"name": "Emily Taylor", "email": "emilytaylor@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
    "7": {"name": "David Lee", "email": "davidlee@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "8": {"name": "Sophia Patel", "email": "sophiapatel@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
    "9": {"name": "Kevin White", "email": "kevinwhite@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "10": {"name": "Olivia Martin", "email": "oliviamartin@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file(filename='attendance.xlsx'):
    try:
        book.save(filename)
        print("Saved!")
    except PermissionError:
        print("Permission denied. Saving with a different name.")
        filename = "attendance_copy.xlsx"
        book.save(filename)
        print(f"Saved as {filename}")

# Add default data to Excel sheet
for roll_no, data in default_data.items():
    sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
    sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
    sheet.cell(row=int(roll_no) + 1, column=3).value = data['email']  # Save email
    for i, attendance in enumerate(data['attendance']):
        sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
print("Default data added to Excel sheet.")

# ... (rest of the code remains the same)

# Add data
def add_data(roll_no, name, email, attendance):
    global sheet

    sheet.cell(row=sheet.max_row + 1, column=1).value = roll_no
    sheet.cell(row=sheet.max_row, column=2).value = name
    sheet.cell(row=sheet.max_row, column=3).value = email  # Save email
    for i, attendance_status in enumerate(attendance):
        sheet.cell(row=sheet.max_row, column=7 + i).value = attendance_status

    save_file()

# Add data
roll_no = int(input("Enter roll number to add data: "))
name = input("Enter name: ")
email = input("Enter email: ")
attendance = input("Enter attendance (Present/Absent): ").split()
add_data(roll_no, name, email, attendance)
In this updated code, I added a column to save the email address for each student. I also updated the add_data function to accept an email address and save it to the Excel sheet.
raceback (most recent call last):
File "c:\Users\Hp\Desktop\python\Python projects\code.py", line 5, in <module>
book = openpyxl.load_workbook('attendance1.xlsx')
File "C:\Users\Hp\AppData\Local\Programs\Python\Python38\lib\site-packages\openpyxl\reader\excel.py", line 346, in load_workbook
reader = ExcelReader(filename, read_only, keep_vba,
File "C:\Users\Hp\AppData\Local\Programs\Python\Python38\lib\site-packages\openpyxl\reader\excel.py", line 123, in __init__
self.archive = _validate_archive(fn)
File "C:\Users\Hp\AppData\Local\Programs\Python\Python38\lib\site-packages\openpyxl\reader\excel.py", line 95, in _validate_archive
archive = ZipFile(filename, 'r')
File "C:\Users\Hp\AppData\Local\Programs\Python\Python38\lib\zipfile.py", line 1251, in __init__
self.fp = io.open(file, filemode)
FileNotFoundError: [Errno 2] No such file or directory: 'attendance1.xlsx'
The error message indicates that the file 'attendance1.xlsx' does not exist in the current working directory.
To fix this issue, you need to make sure that the file 'attendance1.xlsx' exists in the same directory as your Python script. If the file exists in a different directory, you need to specify the full path to the file.
Here's an example:
Python
book = openpyxl.load_workbook('C:\\Users\\Hp\\Desktop\\attendance1.xlsx')
Replace 'C:\\Users\\Hp\\Desktop\\attendance1.xlsx' with the actual path to your file.
Alternatively, you can create the file if it does not exist:
Python
from openpyxl import Workbook

try:
    book = openpyxl.load_workbook('attendance1.xlsx')
except FileNotFoundError:
    book = Workbook()
    book.save('attendance1.xlsx')
In this code, if the file 'attendance1.xlsx' does not exist, a new workbook will be created and saved with that name.
import openpyxl
import os
from openpyxl import Workbook

try:
book = openpyxl.load_workbook('attendance1.xlsx')
except FileNotFoundError:
book = Workbook()
book.save('attendance1.xlsx')
# Load the Excel file
book = openpyxl.load_workbook('attendance1.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
"1": {"name": "John Doe", "email": "johndoe@gmail.com", "attendance": ["Present", "Absent", "Present"]},
"2": {"name": "Jane Smith", "email": "janesmith@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
"3": {"name": "Bob Johnson", "email": "bobjohnson@gmail.com", "attendance": ["Present", "Present", "Absent"]},
"4": {"name": "Alice Brown", "email": "alicebrown@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
"5": {"name": "Mike Davis", "email": "mikedavis@gmail.com", "attendance": ["Present", "Absent", "Present"]},
"6": {"name": "Emily Taylor", "email": "emilytaylor@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
"7": {"name": "David Lee", "email": "davidlee@gmail.com", "attendance": ["Present", "Absent", "Present"]},
"8": {"name": "Sophia Patel", "email": "sophiapatel@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
"9": {"name": "Kevin White", "email": "kevinwhite@gmail.com", "attendance": ["Present", "Absent", "Present"]},
"10": {"name": "Olivia Martin", "email": "oliviamartin@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
}
# Function to save the Excel file
def save_file(filename='attendance.xlsx'):
try:
book.save(filename)
print("Saved!")
except PermissionError:
print("Permission denied. Saving with a different name.")
filename = "attendance_copy.xlsx"
book.save(filename)
print(f"Saved as {filename}")

# Add default data to Excel sheet
for roll_no, data in default_data.items():
sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
sheet.cell(row=int(roll_no) + 1, column=3).value = data['email']
for i, attendance in enumerate(data['attendance']):
sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
print("Default data added to Excel sheet.")

# Function to check attendance
def check(no_of_days, row_num, subject):
global sheet

for i in range(len(row_num)):
if no_of_days[i] == 2:
print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
elif no_of_days[i] > 2:
print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Function to fetch data
def fetch_data(roll_no):
global sheet

for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
return {
"name": sheet.cell(row=i, column=2).value,
"attendance": [sheet.cell(row=i, column=j).value for j in range(7, 10)]
}
return None

# Function to add data
def add_data(roll_no, name, attendance):
global sheet

sheet.cell(row=sheet.max_row + 1, column=1).value = roll_no
sheet.cell(row=sheet.max_row, column=2).value = name
sheet.cell(row=sheet.max_row, column=3).value = email
for i, attendance_status in enumerate(attendance):
sheet.cell(row=sheet.max_row, column=7 + i).value = attendance_status

save_file()

# Function to remove data
def remove_data(roll_no):
global sheet

for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
sheet.delete_rows(i)
save_file()
return
print("Roll number not found.")

# Main loop
while True:
print("1--->CI\n2--->Python\n3--->DM")
subject_choice = int(input("Enter subject: "))

no_of_absentees = int(input('No. of absentees: '))

if no_of_absentees > 1:
roll_nos = list(map(int, input('Roll nos: ').split(' ')))
else:
roll_nos = [int(input('Roll no: '))]

row_nos = []
no_of_days_list = []

for roll_no in roll_nos:
for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
if subject_choice == 1:
no_of_days = sheet.cell(row=i, column=3).value + 1
sheet.cell(row=i, column=3).value = no_of_days
elif subject_choice == 2:
no_of_days = sheet.cell(row=i, column=4).value + 1
sheet.cell(row=i, column=4).value = no_of_days
elif subject_choice == 3:
no_of_days = sheet.cell(row=i, column=5).value + 1
sheet.cell(row=i, column=5).value = no_of_days

row_nos.append(i)
no_of_days_list.append(no_of_days)

save_file()
check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])

cont = input('Another subject? (y/n): ')
if cont.lower() != 'y':
break

# Fetch data
roll_no = int(input("Enter roll number to fetch data: "))
data = fetch_data(roll_no)
if data:
print("Name:", data['name'])
print("Attendance:", data['attendance'])
else:
print("Roll number not found.")

# Add data
roll_no = int(input("Enter roll number to add data: "))
name = input("Enter name: ")
email = input("Enter email: ")
attendance = input("Enter attendance (Present/Absent): ").split()
add_data(roll_no, name, email, attendance)

# Remove data
roll_no = int(input("Enter roll number to remove data: "))
remove_data(roll_no) where is problem Traceback (most recent call last):
File "c:\Users\Hp\Desktop\python\Python projects\tempCodeRunnerFile.py", line 2, in <module>
book = openpyxl.load_workbook('attendance1.xlsx')
NameError: name 'openpyxl' is not defined
The error message indicates that the name 'openpyxl' is not defined. This is because you imported openpyxl with import openpyxl, but you are trying to use it as if you imported it with from openpyxl import *.
To fix this issue, you need to either import openpyxl with from openpyxl import * or use the openpyxl module with its name, like this: openpyxl.load_workbook('attendance1.xlsx').
Here's the corrected code:
Python
from openpyxl import Workbook, load_workbook

try:
    book = load_workbook('attendance1.xlsx')
except FileNotFoundError:
    book = Workbook()
    book.save('attendance1.xlsx')

# Load the Excel file
book = load_workbook('attendance1.xlsx')
sheet = book['Sheet1']

# ... (rest of the code remains the same)
Alternatively, you can keep the original import statement and use the openpyxl module with its name:
Python
import openpyxl

try:
    book = openpyxl.load_workbook('attendance1.xlsx')
except FileNotFoundError:
    book = openpyxl.Workbook()
    book.save('attendance1.xlsx')

# Load the Excel file
book = openpyxl.load_workbook('attendance1.xlsx')
sheet = book['Sheet1']

# ... (rest of the code remains the same)
import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
"1": {"name": "John Doe", "attendance": ["Present", "Absent", "Present"]},
"2": {"name": "Jane Smith", "attendance": ["Absent", "Present", "Absent"]},
"3": {"name": "Bob Johnson", "attendance": ["Present", "Present", "Absent"]},
"4": {"name": "Alice Brown", "attendance": ["Absent", "Absent", "Present"]},
"5": {"name": "Mike Davis", "attendance": ["Present", "Absent", "Present"]},
"6": {"name": "Emily Taylor", "attendance": ["Absent", "Present", "Absent"]},
"7": {"name": "David Lee", "attendance": ["Present", "Absent", "Present"]},
"8": {"name": "Sophia Patel", "attendance": ["Absent", "Absent", "Present"]},
"9": {"name": "Kevin White", "attendance": ["Present", "Absent", "Present"]},
"10": {"name": "Olivia Martin", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
book.save('attendance.xlsx')
print("Saved!")

# Function to check attendance
def check(no_of_days, row_num, subject):
global sheet

for i in range(len(row_num)):
if no_of_days[i] == 2:
print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
elif no_of_days[i] > 2:
print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Function to fetch data
def fetch_data(roll_no):
global sheet

for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
return {
"name": sheet.cell(row=i, column=2).value,
"attendance": [sheet.cell(row=i, column=j).value for j in range(7, 10)]
}
return None

# Function to add data
def add_data(roll_no, name, attendance):
global sheet

sheet.cell(row=sheet.max_row + 1, column=1).value = roll_no
sheet.cell(row=sheet.max_row, column=2).value = name
for i, attendance_status in enumerate(attendance):
sheet.cell(row=sheet.max_row, column=7 + i).value = attendance_status

save_file()

# Function to remove data
def remove_data(roll_no):
global sheet

for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
sheet.delete_rows(i)
save_file()
return
print("Roll number not found.")

# Main loop
while True:
print("1--->CI\n2--->Python\n3--->DM")
subject_choice = int(input("Enter subject: "))

no_of_absentees = int(input('No. of absentees: '))

if no_of_absentees > 1:
roll_nos = list(map(int, input('Roll nos: ').split(' ')))
else:
roll_nos = [int(input('Roll no: '))]

row_nos = []
no_of_days_list = []

for roll_no in roll_nos:
for i in range(2, sheet.max_row + 1):
if sheet.cell(row=i, column=1).value == roll_no:
if subject_choice == 1:
no_of_days = sheet.cell(row=i, column=3).value + 1
sheet.cell(row=i, column=3).value = no_of_days
elif subject_choice == 2:
no_of_days = sheet.cell(row=i, column=4).value + 1
sheet.cell(row=i, column=4).value = no_of_days
elif subject_choice == 3:
no_of_days = sheet.cell(row=i, column=5).value + 1
sheet.cell(row=i, column=5).value = no_of_days

row_nos.append(i)
no_of_days_list.append(no_of_days)

save_file()
check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])
cont = input('Another subject? (y/n): ')
if cont.lower() != 'y':
break

# Add default data to Excel sheet
for roll_no, data in default_data.items():
sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
for i, attendance in enumerate(data['attendance']):
sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
print("Default data added to Excel sheet.")

# Fetch data
roll_no = int(input("Enter roll number to fetch data: "))
data = fetch_data(roll_no)
if data:
print("Name:", data['name'])
print("Attendance:", data['attendance'])
else:
print("Roll number not found.")

# Add data
roll_no = int(input("Enter roll number to add data: "))
name = input("Enter name: ")
attendance = input("Enter attendance (Present/Absent): ").split()
add_data(roll_no, name, attendance)

# Remove data
roll_no = int(input("Enter roll number to remove data: "))
remove_data(roll_no)

# Attendance tracker
# students = {
# "1": {"name": "John Doe", "attendance": ["Present", "Absent", "Present"]},
# "2": {"name": "Jane Smith", "attendance": ["Absent", "Present", "Absent"]},
# "3": {"name": "Bob Johnson", "attendance": ["Present", "Present", "Absent"]},
# "4": {"name": "Alice Brown", "attendance": ["Absent", "Absent", "Present"]},
# "5": {"name": "Mike Davis", "attendance": ["Present", "Absent", "Present"]},
# "6": {"name": "Emily Taylor", "attendance": ["Absent", "Present", "Absent"]},
# "7": {"name": "David Lee", "attendance": ["Present", "Absent", "Present"]},
# "8": {"name": "Sophia Patel", "attendance": ["Absent", "Absent", "Present"]},
# "9": {"name": "Kevin White", "attendance": ["Present", "Absent", "Present"]},
# "10": {"name": "Olivia Martin", "attendance": ["Absent", "Present", "Absent"]},
# }

# def attendance_tracker():
# while True:
# print("\n1. Mark Attendance")
# print("2. View All Students")
# print("3. Search Student")
# print("4. Exit")
# choice = input("Choose an option: ")

# if choice == "1":
# name = input("Enter student name: ")
# roll_no = input("Enter roll number: ")
# status = input("Enter attendance status (Present/Absent): ")
# if roll_no in students:
# students[roll_no]['attendance'].append(status)
# else:
# students[roll_no] = {'name': name, 'attendance': [status]}
# print("Attendance marked successfully!")

# elif choice == "2":
# print("All Students:")
# print("| Roll No | Name | Attendance |")
# print("|---------|-----------------|--------------------|")
# for roll_no, student in students.items():
# print(f"| {roll_no:>7} | {student['name']:>15} | {', '.join(student['attendance']):>20} |")
# print("|---------|-----------------|--------------------|")

# elif choice == "3":
# roll_no = input("Enter roll number to search: ")
# if roll_no in students:
# print(f"Roll No: {roll_no}, Name: {students[roll_no]['name']}, Attendance: {students[roll_no]['attendance']}")
# else:
# print("Student not found!")

# elif choice == "4":
# break

# else:
# print("Invalid option. Please choose a valid option.")

# attendance_tracker() only display name , attenedence and email
Here's the updated code to display only name, attendance, and email:
Python
import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
    "1": {"name": "John Doe", "email": "johndoe@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "2": {"name": "Jane Smith", "email": "janesmith@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
    "3": {"name": "Bob Johnson", "email": "bobjohnson@gmail.com", "attendance": ["Present", "Present", "Absent"]},
    "4": {"name": "Alice Brown", "email": "alicebrown@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
    "5": {"name": "Mike Davis", "email": "mikedavis@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "6": {"name": "Emily Taylor", "email": "emilytaylor@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
    "7": {"name": "David Lee", "email": "davidlee@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "8": {"name": "Sophia Patel", "email": "sophiapatel@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
    "9": {"name": "Kevin White", "email": "kevinwhite@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "10": {"name": "Olivia Martin", "email": "oliviamartin@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
    book.save('attendance.xlsx')
    print("Saved!")

# Function to fetch data
def fetch_data(roll_no):
    global sheet

    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == roll_no:
            return {
                "name": sheet.cell(row=i, column=2).value,
                "email": sheet.cell(row=i, column=3).value,
                "attendance": [sheet.cell(row=i, column=j).value for j in range(7, 10)]
            }
    return None

# Add default data to Excel sheet
for roll_no, data in default_data.items():
    sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
    sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
    sheet.cell(row=int(roll_no) + 1, column=3).value = data['email']
    for i, attendance in enumerate(data['attendance']):
        sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
print("Default data added to Excel sheet.")

# Fetch data
roll_no = int(input("Enter roll number to fetch data: "))
data = fetch_data(roll_no)
if data:
    print("Name:", data['name'])
    print("Email:", data['email'])
    print("Attendance:", data['attendance'])
else:
    print("Roll number not found.")
In this updated code, I have added email to the default data and updated the fetch_data function to return email along with name and attendance. I have also updated the print statements to display email.