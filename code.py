import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
    "1": {"name": "John Doe", "email": "johndoe@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "2": {"name": "Jane Smith", "email": "janesmith@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
    "3": {"name": "Bob Johnson", "email": "bobjohnson@gmail.com", "attendance": ["Present", "Present", "Absent"]},
    "4": {"name": "Alice Brown", "email": "alicebrown@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
    "5": {"name": "Mike Davis", "email": "mikedavis@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "6": {"name": "Emily Taylor", "email": "emilytaylor@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
    "7": {"name": "David Lee", "email": "davidlee@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "8": {"name": "Sophia Patel", "email": "sophiapatel@gmail.com", "attendance": ["Absent", "Absent", "Present"]},
    "9": {"name": "Kevin White", "email": "kevinwhite@gmail.com", "attendance": ["Present", "Absent", "Present"]},
    "10": {"name": "Olivia Martin", "email": "oliviamartin@gmail.com", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
    book.save('attendance.xlsx')
    print("Saved!")

# Function to fetch data
def fetch_data(roll_no):
    global sheet

    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == roll_no:
            return {
                "name": sheet.cell(row=i, column=2).value,
                "email": sheet.cell(row=i, column=3).value,
                "attendance": [sheet.cell(row=i, column=j).value for j in range(7, 10)]
            }
    return None

# Add default data to Excel sheet
for roll_no, data in default_data.items():
    sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
    sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
    sheet.cell(row=int(roll_no) + 1, column=3).value = data['email']
    for i, attendance in enumerate(data['attendance']):
        sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
print("Default data added to Excel sheet.")

# Fetch data
roll_no = int(input("Enter roll number to fetch data: "))
data = fetch_data(roll_no)
if data:
    print("Name:", data['name'])
    print("Email:", data['email'])
    print("Attendance:", data['attendance'])
else:
    print("Roll number not found.")