import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Function to save the Excel file
def save_file():
    book.save('attendance.xlsx')
    print("Saved!")

# Function to check attendance
def check(no_of_days, row_num, subject):
    global sheet

    for i in range(len(row_num)):
        if no_of_days[i] == 2:
            print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
        elif no_of_days[i] > 2:
            print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Main loop
while True:
    print("1--->CI\n2--->Python\n3--->DM")
    subject_choice = int(input("Enter subject: "))

    no_of_absentees = int(input('No. of absentees: '))

    if no_of_absentees > 1:
        roll_nos = list(map(int, input('Roll nos: ').split(' ')))
    else:
        roll_nos = [int(input('Roll no: '))]

    row_nos = []
    no_of_days_list = []

    for roll_no in roll_nos:
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == roll_no:
                if subject_choice == 1:
                    no_of_days = sheet.cell(row=i, column=3).value + 1
                    sheet.cell(row=i, column=3).value = no_of_days
                elif subject_choice == 2:
                    no_of_days = sheet.cell(row=i, column=4).value + 1
                    sheet.cell(row=i, column=4).value = no_of_days
                elif subject_choice == 3:
                    no_of_days = sheet.cell(row=i, column=5).value + 1
                    sheet.cell(row=i, column=5).value = no_of_days

                row_nos.append(i)
                no_of_days_list.append(no_of_days)

    save_file()
    check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])

    cont = input('Another subject? (y/n): ')
    if cont.lower() != 'y':
        break