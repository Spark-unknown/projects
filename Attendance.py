import openpyxl

# Load the Excel file
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Default data
default_data = {
    "1": {"name": "John Doe", "attendance": ["Present", "Absent", "Present"]},
    "2": {"name": "Jane Smith", "attendance": ["Absent", "Present", "Absent"]},
    "3": {"name": "Bob Johnson", "attendance": ["Present", "Present", "Absent"]},
    "4": {"name": "Alice Brown", "attendance": ["Absent", "Absent", "Present"]},
    "5": {"name": "Mike Davis", "attendance": ["Present", "Absent", "Present"]},
    "6": {"name": "Emily Taylor", "attendance": ["Absent", "Present", "Absent"]},
    "7": {"name": "David Lee", "attendance": ["Present", "Absent", "Present"]},
    "8": {"name": "Sophia Patel", "attendance": ["Absent", "Absent", "Present"]},
    "9": {"name": "Kevin White", "attendance": ["Present", "Absent", "Present"]},
    "10": {"name": "Olivia Martin", "attendance": ["Absent", "Present", "Absent"]},
}

# Function to save the Excel file
def save_file():
    book.save('attendance.xlsx')
    print("Saved!")

# Function to check attendance
def check(no_of_days, row_num, subject):
    global sheet

    for i in range(len(row_num)):
        if no_of_days[i] == 2:
            print(f"Warning: Student {sheet.cell(row=row_num[i], column=1).value} has 2 days of absence in {subject}")
        elif no_of_days[i] > 2:
            print(f"Student {sheet.cell(row=row_num[i], column=1).value} has more than 2 days of absence in {subject}")

# Function to fetch data
def fetch_data(roll_no):
    global sheet

    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == roll_no:
            return {
                "name": sheet.cell(row=i, column=2).value,
                "attendance": [sheet.cell(row=i, column=j).value for j in range(7, 10)]
            }
    return None

# Function to add data
def add_data(roll_no, name, attendance):
    global sheet

    sheet.cell(row=sheet.max_row + 1, column=1).value = roll_no
    sheet.cell(row=sheet.max_row, column=2).value = name
    for i, attendance_status in enumerate(attendance):
        sheet.cell(row=sheet.max_row, column=7 + i).value = attendance_status

    save_file()

# Function to remove data
def remove_data(roll_no):
    global sheet

    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == roll_no:
            sheet.delete_rows(i)
            save_file()
            return
    print("Roll number not found.")

# Main loop
while True:
    print("1--->CI\n2--->Python\n3--->DM")
    subject_choice = int(input("Enter subject: "))

    no_of_absentees = int(input('No. of absentees: '))

    if no_of_absentees > 1:
        roll_nos = list(map(int, input('Roll nos: ').split(' ')))
    else:
        roll_nos = [int(input('Roll no: '))]

    row_nos = []
    no_of_days_list = []

    for roll_no in roll_nos:
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == roll_no:
                if subject_choice == 1:
                    no_of_days = sheet.cell(row=i, column=3).value + 1
                    sheet.cell(row=i, column=3).value = no_of_days
                elif subject_choice == 2:
                    no_of_days = sheet.cell(row=i, column=4).value + 1
                    sheet.cell(row=i, column=4).value = no_of_days
                elif subject_choice == 3:
                    no_of_days = sheet.cell(row=i, column=5).value + 1
                    sheet.cell(row=i, column=5).value = no_of_days

                row_nos.append(i)
                no_of_days_list.append(no_of_days)

    save_file()
    check(no_of_days_list, row_nos, ['CI', 'Python', 'DM'][subject_choice - 1])
    cont = input('Another subject? (y/n): ')
    if cont.lower() != 'y':
        break

# Add default data to Excel sheet
for roll_no, data in default_data.items():
    sheet.cell(row=int(roll_no) + 1, column=1).value = int(roll_no)
    sheet.cell(row=int(roll_no) + 1, column=2).value = data['name']
    for i, attendance in enumerate(data['attendance']):
        sheet.cell(row=int(roll_no) + 1, column=7 + i).value = attendance

save_file()
print("Default data added to Excel sheet.")

# Fetch data
roll_no = int(input("Enter roll number to fetch data: "))
data = fetch_data(roll_no)
if data:
    print("Name:", data['name'])
    print("Attendance:", data['attendance'])
else:
    print("Roll number not found.")

# Add data
roll_no = int(input("Enter roll number to add data: "))
name = input("Enter name: ")
attendance = input("Enter attendance (Present/Absent): ").split()
add_data(roll_no, name, attendance)

# Remove data
roll_no = int(input("Enter roll number to remove data: "))
remove_data(roll_no)

 # Attendance tracker
# students = {
#     "1": {"name": "John Doe", "attendance": ["Present", "Absent", "Present"]},
#     "2": {"name": "Jane Smith", "attendance": ["Absent", "Present", "Absent"]},
#     "3": {"name": "Bob Johnson", "attendance": ["Present", "Present", "Absent"]},
#     "4": {"name": "Alice Brown", "attendance": ["Absent", "Absent", "Present"]},
#     "5": {"name": "Mike Davis", "attendance": ["Present", "Absent", "Present"]},
#     "6": {"name": "Emily Taylor", "attendance": ["Absent", "Present", "Absent"]},
#     "7": {"name": "David Lee", "attendance": ["Present", "Absent", "Present"]},
#     "8": {"name": "Sophia Patel", "attendance": ["Absent", "Absent", "Present"]},
#     "9": {"name": "Kevin White", "attendance": ["Present", "Absent", "Present"]},
#     "10": {"name": "Olivia Martin", "attendance": ["Absent", "Present", "Absent"]},
# }

# def attendance_tracker():
#     while True:
#         print("\n1. Mark Attendance")
#         print("2. View All Students")
#         print("3. Search Student")
#         print("4. Exit")
#         choice = input("Choose an option: ")

#         if choice == "1":
#             name = input("Enter student name: ")
#             roll_no = input("Enter roll number: ")
#             status = input("Enter attendance status (Present/Absent): ")
#             if roll_no in students:
#                 students[roll_no]['attendance'].append(status)
#             else:
#                 students[roll_no] = {'name': name, 'attendance': [status]}
#             print("Attendance marked successfully!")

#         elif choice == "2":
#             print("All Students:")
#             print("| Roll No | Name            | Attendance         |")
#             print("|---------|-----------------|--------------------|")
#             for roll_no, student in students.items():
#                 print(f"| {roll_no:>7} | {student['name']:>15} | {', '.join(student['attendance']):>20} |")
#             print("|---------|-----------------|--------------------|")

#         elif choice == "3":
#             roll_no = input("Enter roll number to search: ")
#             if roll_no in students:
#                 print(f"Roll No: {roll_no}, Name: {students[roll_no]['name']}, Attendance: {students[roll_no]['attendance']}")
#             else:
#                 print("Student not found!")

#         elif choice == "4":
#             break

#         else:
#             print("Invalid option. Please choose a valid option.")

# attendance_tracker()